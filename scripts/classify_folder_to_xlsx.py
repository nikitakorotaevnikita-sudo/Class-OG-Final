"""Инструмент: классифицировать все .txt в папке через работающий сервис и
сохранить результаты в xlsx в «референс-формате» (как файлы 16-выборки).

Формат ячейки результата (одна строка = одно обращение):
    [⚑ СРАБОТАЛ FALLBACK]           <- только если сработал full-classifier fallback
    [Вид: … · Тип: … · общая уверенность: …]

    Вопрос                          <- «Вопрос 1/2/…» если вопросов больше одного
    Код: …
    Тема: …
    Путь: …
    Уверенность: …
    Пояснение: …

Использование (запускать через PowerShell — корректно передаёт кириллические пути):
    python scripts/classify_folder_to_xlsx.py "<папка с .txt>" "<путь к .xlsx>" [url] [llm_provider]
    # url по умолчанию http://127.0.0.1:8010/classify (прод); llm_provider опционален.

Не читает и не пересказывает тексты обращений — только прогоняет через пайплайн.
"""
import sys, os, glob, time, json
import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

DIR = sys.argv[1]
OUT = sys.argv[2]
URL = sys.argv[3] if len(sys.argv) > 3 else "http://127.0.0.1:8010/classify"
PROVIDER = sys.argv[4] if len(sys.argv) > 4 else None

files = sorted(glob.glob(os.path.join(DIR, "*.txt")))
print(f"файлов: {len(files)} | сервис: {URL} | provider: {PROVIDER or 'по умолчанию'}", flush=True)


def classify(text):
    payload = {"appeal_text": text}
    if PROVIDER:
        payload["llm_provider"] = PROVIDER
    last = None
    for _ in range(3):
        try:
            r = httpx.post(URL, json=payload, timeout=240)
            d = r.json()
            if d.get("questions"):
                return d
            last = f"HTTP {r.status_code}: {str(d)[:120]}"
        except Exception as e:
            last = f"{type(e).__name__}: {e}"
        time.sleep(5)
    return {"_error": last}


results = []
for i, path in enumerate(files, 1):
    name = os.path.basename(path)
    text = open(path, encoding="utf-8", errors="replace").read().strip()
    t0 = time.time()
    d = classify(text) if text else {"_error": "пустой файл"}
    if d.get("_error"):
        results.append({"file": name, "text": text, "error": d["_error"]})
        print(f"[{i}/{len(files)}] {name}: ОШИБКА {d['_error'][:80]}", flush=True)
        continue
    results.append({
        "file": name, "text": text,
        "vid": d.get("vid_obrascheniya"), "tip": d.get("tip_obrascheniya"),
        "overall_confidence": d.get("overall_confidence"),
        "needs_verification": d.get("needs_verification"),
        "full_fallback_used": d.get("full_fallback_used"),
        "questions": [{"code": q.get("code"), "name": q.get("name"), "full_path": q.get("full_path"),
                       "confidence": q.get("confidence"), "reasoning": q.get("reasoning")}
                      for q in d.get("questions", [])],
    })
    print(f"[{i}/{len(files)}] {name}: {len(d.get('questions', []))}q "
          f"conf={d.get('overall_confidence')} fb={d.get('full_fallback_used')} {time.time()-t0:.0f}s", flush=True)

# ── xlsx ────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Классификация"
ws.append(["№", "Текст обращения", "Прототип RAG + LLM"])
for c in ws[1]:
    c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9E1F2")
    c.alignment = Alignment(vertical="top", wrap_text=True)

for i, r in enumerate(results, 1):
    if r.get("error"):
        cell = "ОШИБКА: " + r["error"]
    else:
        qs = r.get("questions", [])
        multi = len(qs) > 1
        blocks = []
        for j, q in enumerate(qs, 1):
            head = f"Вопрос {j}" if multi else "Вопрос"
            conf = q.get("confidence")
            cs = f"{conf:.2f}" if isinstance(conf, (int, float)) else str(conf)
            blocks.append(f"{head}\nКод: {q.get('code')}\nТема: {q.get('name')}\n"
                          f"Путь: {q.get('full_path')}\nУверенность: {cs}\nПояснение: {q.get('reasoning')}")
        fb = "⚑ СРАБОТАЛ FALLBACK\n" if r.get("full_fallback_used") else ""
        meta = f"[Вид: {r.get('vid')} · Тип: {r.get('tip')} · общая уверенность: {r.get('overall_confidence')}]"
        cell = fb + meta + "\n\n" + "\n\n".join(blocks)
    ws.append([i, r.get("text"), cell])

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 95
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
wb.save(OUT)

ok = sum(1 for r in results if not r.get("error"))
fb = sum(1 for r in results if r.get("full_fallback_used"))
nv = sum(1 for r in results if not r.get("error") and r.get("needs_verification"))
print(f"\nСохранено: {OUT}")
print(f"обращений: {len(results)} | успешно: {ok} | fallback: {fb} | на верификацию: {nv}", flush=True)
