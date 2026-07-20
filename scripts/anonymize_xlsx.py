"""Инструмент: обезличивание персональных данных в xlsx через Ario (Qwen3.6).

Заменяет ПДн на нейтральные плейсхолдеры:
  ФИО физлиц        → «Иванов Иван Иванович»
  адрес             → «г. Город, ул. Улица, д. 1, кв. 1»
  телефон           → «0000000000»
  email             → «example@example.com»
  СНИЛС/ЕСИА/ИНН/паспорт/№ счёта → нули
  дата рождения     → «01.01.1970»
НЕ трогает: суть обращения, коды классификатора (XXXX.XXXX.XXXX...), названия
гос.органов/должностей/тематику, структуру ячейки.

Обрабатываются колонки, чьи заголовки содержат подстроки из TARGET_HEADERS
(по умолч. «Текст обращения» и «Результат»). Остальные колонки копируются как есть.

Использование (через PowerShell — кириллические пути):
    python scripts/anonymize_xlsx.py "<вход.xlsx>" "<выход.xlsx>"

При ошибке LLM ячейка помечается «[⚠ НЕ ОБЕЗЛИЧЕНО — проверить вручную]»
(исходный текст с ПДн НЕ выводится, чтобы не допустить утечку).
"""
import sys, time
import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, "src")
sys.stdout.reconfigure(encoding="utf-8")
from config import ARIO_API_KEY, ARIO_BASE_URL, ARIO_MODEL

IN = sys.argv[1]
OUT = sys.argv[2]
TARGET_HEADERS = ("текст обращения", "результат")

SYSTEM = (
    "Ты — система обезличивания персональных данных (анонимизация) в текстах обращений граждан. "
    "Верни ТОТ ЖЕ текст, заменив ВСЕ персональные данные на нейтральные плейсхолдеры:\n"
    "• ФИО любых физических лиц (заявитель, родственники, любые люди) → «Иванов Иван Иванович» "
    "(или «Иванова Ивана Ивановна» для явно женского; инициалы → «И.И.»);\n"
    "• почтовый адрес (город, улица, дом, квартира) → «г. Город, ул. Улица, д. 1, кв. 1»;\n"
    "• номер телефона → «0000000000»;\n"
    "• email → «example@example.com»;\n"
    "• СНИЛС, ЕСИА, ИНН, серия/номер паспорта, номер счёта, кадастровый номер → нули соответствующей длины;\n"
    "• дата рождения → «01.01.1970».\n"
    "СТРОГО СОХРАНИ без изменений: суть и смысл обращения; коды классификатора формата "
    "XXXX.XXXX.XXXX.XXXX; названия гос.органов, ведомств, должностей (Губернатор, Глава ЦБ и т.п.); "
    "тематику; структуру и разметку текста (строки «Код:», «Тема:», «Путь:», «Пояснение:» и т.п.). "
    "Не удаляй и не добавляй строки. Верни ТОЛЬКО обезличенный текст, без комментариев и пояснений."
)


def anonymize(text: str) -> str | None:
    payload = {
        "model": ARIO_MODEL,
        "messages": [{"role": "system", "content": SYSTEM},
                     {"role": "user", "content": text}],
        "temperature": 0,
        "max_tokens": 4000,
    }
    for _ in range(3):
        try:
            r = httpx.post(f"{ARIO_BASE_URL}/chat/completions",
                           headers={"Authorization": f"Bearer {ARIO_API_KEY}", "Content-Type": "application/json"},
                           json=payload, timeout=180)
            if r.status_code == 200:
                out = r.json()["choices"][0]["message"]["content"].strip()
                for pref in ("```", "```text"):
                    if out.startswith(pref):
                        out = out[len(pref):].strip()
                if out.endswith("```"):
                    out = out[:-3].strip()
                return out
        except Exception:
            pass
        time.sleep(4)
    return None


wb = openpyxl.load_workbook(IN)
ws = wb.active
headers = [str(c.value or "").lower() for c in ws[1]]
target_cols = [i + 1 for i, h in enumerate(headers) if any(t in h for t in TARGET_HEADERS)]
print(f"колонки для обезличивания: {[ws.cell(1, c).value for c in target_cols]}", flush=True)

total = fail = 0
for row in range(2, ws.max_row + 1):
    for col in target_cols:
        cell = ws.cell(row, col)
        val = cell.value
        if not val or not str(val).strip():
            continue
        total += 1
        res = anonymize(str(val))
        if res is None:
            fail += 1
            cell.value = "[⚠ НЕ ОБЕЗЛИЧЕНО — ошибка LLM, проверить вручную]"
        else:
            cell.value = res
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if row % 5 == 0:
        print(f"  строк обработано: {row-1}/{ws.max_row-1}", flush=True)
        wb.save(OUT)

wb.save(OUT)
print(f"\nСохранено: {OUT}")
print(f"ячеек обезличено: {total-fail}/{total} | ошибок LLM: {fail}", flush=True)
